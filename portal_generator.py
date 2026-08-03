#!/usr/bin/env python3
"""
DRS Owner Portal generator — rebuilt 2026-07-31 ("start over" spec).

Reads turo.xlsx (export of live Google Sheet "Turo Master",
id 19iwi1OhIzDTAhKhHjaPeQzWTZPVMaIiB2cy4DkMaO38).

Rules (Lindsey, 2026-07-31):
- Only statement months >= 6/2026 are used, capped at the current statement month
  (no future/upcoming months are listed, for statements or expenses).
- Owner statements come from OwnerStatements columns A:I
  (Statement Month, Owner, Vehicle, Days, Revenue, Management Fee,
   Owner Adjustments, Operating Expenses, Client Payout).
- Expenses come from ExpenseMaster columns A:H.
  Column F (Amount): negative = CREDIT to the owner, positive = deduction.
- Owner/vehicle overrides (below) are applied everywhere.
- Upcoming bookings calendar: ReservationsMaster rows with
  Trip status == 'Booked' and Trip end >= today; the ONLY dollar figure shown
  is Revenue Subject to Split (col index 51).

Usage: python3 portal_generator.py <owner name or email> [--out FILE] [--today YYYY-MM-DD]
       python3 portal_generator.py --snapshot   (print per-owner change-detection state JSON)
"""
import sys, json, hashlib, datetime, html, calendar as callib
import openpyxl

XLSX = 'turo.xlsx'
DATA_START = (2026, 6)   # first statement month used

# ---- overrides directed by Lindsey 2026-07-31 -------------------------------
OWNER_OVERRIDES = {           # by VIN (uppercased)
    '1GNSKHKC1LR274922': 'Jayson Cahill',
    '1GNSKHKC3LR187573': 'Jayson Cahill',
    '1C4RJHET1N8536935': 'Joel Bowers',
    '5LMJJ3LT2KEL19987': 'Zachary Holderness',
}
VEHICLE_OVERRIDES = {         # by VIN (uppercased)
    '1GNSKHKC1LR274922': 'Black 2020 Chevrolet Suburban',
    '1GNSKHKC3LR187573': 'White 2020 Chevrolet Suburban',
    '1GNSKHKC1LR177317': 'Blue 2020 Chevrolet Suburban',
}
NAME_FIXES = {'Jason Cahill': 'Jayson Cahill', 'Zach Holderness': 'Zachary Holderness'}

def vin_key(v): return str(v or '').strip().upper()

def fix_owner(name, vin):
    n = str(name or '').strip()
    n = NAME_FIXES.get(n, n)
    return OWNER_OVERRIDES.get(vin_key(vin), n)

def fix_vehicle(name, vin):
    return VEHICLE_OVERRIDES.get(vin_key(vin), str(name or '').strip())

def mkey(dt):
    if isinstance(dt, datetime.datetime) or isinstance(dt, datetime.date):
        return (dt.year, dt.month)
    return None

def mlabel(k): return f'{k[1]}/{k[0]}'

def num(v):
    try: return float(v)
    except (TypeError, ValueError): return 0.0

def money(v, cents=True):
    neg = v < -1e-9
    s = f'${abs(v):,.2f}' if cents else f'${abs(v):,.0f}'
    return ('-' + s) if neg else s

# ---- load -------------------------------------------------------------------
def load(today):
    cur = (today.year, today.month)
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    # Owners roster: split + display info per VIN
    roster = {}
    ws = wb['Owners']; it = ws.iter_rows(values_only=True); hdr = [str(h) for h in next(it)]
    ix = {h: i for i, h in enumerate(hdr)}
    for r in it:
        if not r or not r[ix['VIN']]: continue
        vin = vin_key(r[ix['VIN']])
        owner = fix_owner(f"{r[ix['First']]} {r[ix['Last']]}", vin)
        roster[vin] = {
            'owner': owner,
            'email': str(r[ix['Owner Email']] or '').strip().lower(),
            'split': num(r[ix['Rev Split']]) or 0.7,
            'desc': f"{r[ix['Color']]} {str(r[ix['Year']]).replace('.0','')} {r[ix['Make']]} {r[ix['Model']]}",
        }

    # OwnerStatements A:I, months in [DATA_START, current]
    stmts = {}   # vin -> [rows]
    ws = wb['OwnerStatements']; it = ws.iter_rows(values_only=True); next(it)
    for r in it:
        if not r or r[0] is None: continue
        k = mkey(r[0])
        if not k or k < DATA_START or k > cur: continue
        vin = vin_key(r[9])
        stmts.setdefault(vin, []).append({
            'month': k, 'owner': fix_owner(r[1], vin), 'vehicle': fix_vehicle(r[2], vin),
            'days': num(r[3]), 'revenue': num(r[4]), 'mgmt_fee': num(r[5]),
            'adjustments': num(r[6]), 'op_expenses': num(r[7]), 'payout': num(r[8]),
        })

    # ExpenseMaster A:H, months in [DATA_START, current] — no upcoming months
    exps = {}
    ws = wb['ExpenseMaster']; it = ws.iter_rows(values_only=True); next(it)
    for r in it:
        if not r or r[1] is None: continue
        k = mkey(r[1])
        if not k or k < DATA_START or k > cur: continue
        vin = vin_key(r[4])
        amt = num(r[5])
        exps.setdefault(vin, []).append({
            'month': k, 'vendor': str(r[6] or '').strip(), 'desc': str(r[7] or '').strip(),
            'amount': amt, 'credit': amt < 0,
        })

    # Upcoming bookings: Booked & Trip end >= today
    trips = {}
    ws = wb['ReservationsMaster']; it = ws.iter_rows(values_only=True); next(it)
    for r in it:
        if not r or len(r) < 52: continue
        status = str(r[10] or '').strip()
        end = r[7]; start = r[6]
        if status != 'Booked' or not isinstance(end, datetime.datetime) or end < today: continue
        vin = vin_key(r[5])
        trips.setdefault(vin, []).append({
            'guest': str(r[1] or '').strip(), 'start': start, 'end': end, 'rss': num(r[51]),
        })
    for v in trips.values(): v.sort(key=lambda t: t['start'])
    for v in stmts.values(): v.sort(key=lambda s: s['month'])
    for v in exps.values(): v.sort(key=lambda e: (e['month'], e['vendor']))
    return roster, stmts, exps, trips

# ---- snapshot for change detection ------------------------------------------
def snapshot(roster, stmts, exps, trips):
    owners = {}
    for vin, info in roster.items():
        owners.setdefault(info['owner'], []).append(vin)
    out = {}
    for owner, vins in sorted(owners.items()):
        blob = json.dumps({
            v: {'s': stmts.get(v, []), 'e': exps.get(v, []),
                't': [{**t, 'start': t['start'].isoformat(), 'end': t['end'].isoformat()} for t in trips.get(v, [])]}
            for v in sorted(vins)
        }, default=str, sort_keys=True)
        out[owner] = {'vins': sorted(vins), 'hash': hashlib.sha256(blob.encode()).hexdigest()[:16]}
    return out

# ---- calendar rendering -----------------------------------------------------
PALETTE = ['#2563eb', '#059669', '#d97706', '#7c3aed', '#dc2626', '#0891b2']

def render_calendar(trips, today):
    """Month grids from current month through last trip month; trip days shaded,
    labeled with revenue subject to split only."""
    if not trips:
        return '<p class="muted">No upcoming trips currently booked.</p>'
    last = max(t['end'] for t in trips)
    months, y, m = [], today.year, today.month
    while (y, m) <= (last.year, last.month):
        months.append((y, m))
        m += 1
        if m == 13: m, y = 1, y + 1
    day_map = {}  # date -> list of trip idx
    for i, t in enumerate(trips):
        d = t['start'].date()
        while d <= t['end'].date():
            day_map.setdefault(d, []).append(i)
            d += datetime.timedelta(days=1)
    out = ['<div class="calwrap">']
    for (y, m) in months:
        out.append(f'<div class="cal"><div class="calhead">{callib.month_name[m]} {y}</div>')
        out.append('<div class="calgrid">')
        for wd in ['Su','Mo','Tu','We','Th','Fr','Sa']:
            out.append(f'<div class="dow">{wd}</div>')
        first_wd = (callib.weekday(y, m, 1) + 1) % 7  # Sunday-start
        for _ in range(first_wd): out.append('<div class="day empty"></div>')
        for d in range(1, callib.monthrange(y, m)[1] + 1):
            dt = datetime.date(y, m, d)
            ids = day_map.get(dt, [])
            cls, style = 'day', ''
            if ids:
                cls += ' booked'
                style = f'background:{PALETTE[ids[0] % len(PALETTE)]}22;border-color:{PALETTE[ids[0] % len(PALETTE)]}'
            if dt == today.date(): cls += ' today'
            out.append(f'<div class="{cls}" style="{style}">{d}</div>')
        out.append('</div></div>')
    out.append('</div>')
    # legend / trip chips with RSS only
    out.append('<div class="triplegend">')
    for i, t in enumerate(trips):
        c = PALETTE[i % len(PALETTE)]
        out.append(
            f'<span class="chip" style="border-color:{c}"><span class="dot" style="background:{c}"></span>'
            f'{html.escape(t["guest"])} · {t["start"]:%-m/%-d}–{t["end"]:%-m/%-d} · '
            f'<b>{money(t["rss"])}</b> revenue subject to split</span>')
    out.append('</div>')
    return ''.join(out)

# ---- portal rendering -------------------------------------------------------
def render(owner, roster, stmts, exps, trips, today):
    vins = [v for v, i in roster.items() if i['owner'].lower() == owner.lower()
            or i['email'] == owner.lower()]
    if not vins: raise SystemExit(f'No vehicles found for {owner!r}')
    owner_name = roster[vins[0]]['owner']
    cur = (today.year, today.month)
    sections = []
    total_cur_payout = 0.0
    for vin in sorted(vins):
        info = roster[vin]
        s_rows = stmts.get(vin, [])
        e_rows = exps.get(vin, [])
        t_rows = trips.get(vin, [])
        veh = fix_vehicle(s_rows[0]['vehicle'] if s_rows else info['desc'], vin)
        split = info['split']
        cur_row = next((r for r in s_rows if r['month'] == cur), None)
        total_cur_payout += cur_row['payout'] if cur_row else 0.0
        rev = sum(r['revenue'] for r in s_rows)

        # statements table (A:I fields)
        st = ['<table><thead><tr><th>Month</th><th class="r">Days</th><th class="r">Revenue</th>'
              '<th class="r">Mgmt Fee</th><th class="r">Adjustments</th>'
              '<th class="r">Operating Expenses</th><th class="r">Owner Payout</th></tr></thead><tbody>']
        for r in s_rows:
            pc = ' class="r neg"' if r['payout'] < 0 else ' class="r pos"'
            st.append(f"<tr><td>{mlabel(r['month'])}</td><td class='r'>{int(r['days'])}</td>"
                      f"<td class='r'>{money(r['revenue'])}</td><td class='r'>{money(r['mgmt_fee'])}</td>"
                      f"<td class='r'>{money(r['adjustments'])}</td><td class='r'>{money(r['op_expenses'])}</td>"
                      f"<td{pc}><b>{money(r['payout'])}</b></td></tr>")
        st.append('</tbody></table>')

        # expenses table — earliest month first, credits flagged
        if e_rows:
            ded = sum(e['amount'] for e in e_rows if not e['credit'])
            cred = -sum(e['amount'] for e in e_rows if e['credit'])
            ex = ['<table><thead><tr><th>Month</th><th>Vendor</th><th>Description</th>'
                  '<th class="r">Amount</th></tr></thead><tbody>']
            for e in e_rows:
                if e['credit']:
                    amt = f'<span class="pos">{money(-e["amount"])} credit</span>'
                else:
                    amt = money(e['amount'])
                ex.append(f"<tr><td>{mlabel(e['month'])}</td><td>{html.escape(e['vendor'])}</td>"
                          f"<td>{html.escape(e['desc'])}</td><td class='r'>{amt}</td></tr>")
            foot = f"<tr class='total'><td colspan='3'>Total deductions</td><td class='r'>{money(ded)}</td></tr>"
            if cred > 0:
                foot += f"<tr class='total'><td colspan='3'>Total credits</td><td class='r pos'>{money(cred)}</td></tr>"
            ex.append(foot + '</tbody></table>')
            ex = ''.join(ex)
        else:
            ex = '<p class="muted">No expenses recorded for this period.</p>'

        upcoming_rev = sum(t['rss'] for t in t_rows)
        sections.append(f"""
<section class="vehicle">
  <div class="vcard">
    <div><div class="vname">{html.escape(veh)}</div>
    <div class="vmeta">VIN {vin} &nbsp;·&nbsp; Revenue split {int(split*100)}% owner / {int((1-split)*100)}% DRS</div></div>
  </div>
  <div class="kpis">
    <div class="kpi"><div class="klabel">Current statement ({mlabel(cur)})</div>
      <div class="kval {'neg' if (cur_row and cur_row['payout']<0) else ''}">{money(cur_row['payout']) if cur_row else '—'}</div></div>
    <div class="kpi"><div class="klabel">Revenue since 6/2026</div><div class="kval">{money(rev)}</div></div>
    <div class="kpi"><div class="klabel">Upcoming trips</div><div class="kval">{len(t_rows)}</div>
      <div class="ksub">{money(upcoming_rev)} revenue subject to split</div></div>
  </div>
  <h2>Payout Statements</h2>
  <div class="note blue"><b>Notes:</b><br>
  If expenses exceed the payout amount, an invoice will come via email. Negative balances will
  no longer roll over.<br>
  Outstanding invoices (vendors, insurance, etc.) must be paid before receiving monthly payouts.</div>
  {''.join(st)}
  <h2>Expenses</h2>
  {ex}
  <h2>Upcoming Bookings</h2>
  {render_calendar(t_rows, today)}
</section>""")

    return f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{html.escape(owner_name)} — DRS Owner Portal</title>
<style>
:root {{ --ink:#1a2333; --mut:#64748b; --line:#e2e8f0; --bg:#f6f8fb; --card:#fff;
  --brand:#1d4ed8; --pos:#059669; --neg:#dc2626; }}
* {{ box-sizing:border-box; }}
body {{ margin:0; font:15px/1.55 -apple-system,'Segoe UI',Roboto,Helvetica,Arial,sans-serif;
  color:var(--ink); background:var(--bg); }}
.wrap {{ max-width:960px; margin:0 auto; padding:28px 20px 60px; }}
header h1 {{ margin:0 0 2px; font-size:26px; }}
header .sub {{ color:var(--mut); margin-bottom:18px; }}
.note {{ border-radius:10px; padding:12px 16px; margin:14px 0; font-size:14px; }}
.note.amber {{ background:#fef3c7; border:1px solid #f59e0b; }}
.note.blue {{ background:#dbeafe; border:1px solid #3b82f6; }}
section.vehicle {{ background:var(--card); border:1px solid var(--line); border-radius:14px;
  padding:22px 24px; margin:22px 0; box-shadow:0 1px 3px rgba(16,24,40,.06); }}
.vname {{ font-size:20px; font-weight:700; }}
.vmeta {{ color:var(--mut); font-size:13px; }}
.kpis {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(180px,1fr)); gap:12px; margin:18px 0; }}
.kpi {{ background:var(--bg); border:1px solid var(--line); border-radius:10px; padding:12px 14px; }}
.klabel {{ font-size:12px; color:var(--mut); text-transform:uppercase; letter-spacing:.04em; }}
.kval {{ font-size:22px; font-weight:700; margin-top:2px; }}
.ksub {{ font-size:12px; color:var(--mut); }}
h2 {{ font-size:16px; margin:26px 0 8px; }}
table {{ width:100%; border-collapse:collapse; font-size:14px; }}
th,td {{ padding:8px 10px; border-bottom:1px solid var(--line); text-align:left; }}
th {{ font-size:12px; color:var(--mut); text-transform:uppercase; letter-spacing:.03em; }}
td.r, th.r {{ text-align:right; }}
tr.total td {{ font-weight:700; border-top:2px solid var(--ink); }}
.pos {{ color:var(--pos); }} .neg {{ color:var(--neg); }}
.muted {{ color:var(--mut); }}
.fine {{ font-size:12px; color:var(--mut); }}
.calwrap {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(240px,1fr)); gap:16px; }}
.cal {{ border:1px solid var(--line); border-radius:10px; overflow:hidden; }}
.calhead {{ background:var(--ink); color:#fff; text-align:center; padding:6px; font-weight:600; font-size:14px; }}
.calgrid {{ display:grid; grid-template-columns:repeat(7,1fr); gap:2px; padding:6px; }}
.dow {{ font-size:10px; color:var(--mut); text-align:center; }}
.day {{ aspect-ratio:1; display:flex; align-items:center; justify-content:center; font-size:12px;
  border:1px solid transparent; border-radius:6px; }}
.day.booked {{ font-weight:700; }}
.day.today {{ outline:2px solid var(--brand); }}
.triplegend {{ margin-top:12px; display:flex; flex-wrap:wrap; gap:8px; }}
.chip {{ border:1px solid; border-radius:999px; padding:4px 12px; font-size:13px; background:#fff; }}
.dot {{ display:inline-block; width:8px; height:8px; border-radius:50%; margin-right:6px; }}
footer {{ color:var(--mut); font-size:12px; margin-top:26px; text-align:center; }}
@media print {{ body {{ background:#fff; }} section.vehicle {{ box-shadow:none; }} }}
</style></head><body><div class="wrap">
<header><h1>{html.escape(owner_name)}</h1>
<div class="sub">DIA RideSource · Owner Portal · generated {today:%-m/%-d/%Y}</div></header>
<div class="note amber"><b>Manager's note:</b> Welcome to the Owner Portal! This is a work in
progress, but this is a nice starter. My goal is to have this automatically update daily.</div>
{''.join(sections)}
<footer>DIA RideSource, LLC · Questions? diaridesource@gmail.com</footer>
</div></body></html>"""

# ---- main -------------------------------------------------------------------
if __name__ == '__main__':
    args = sys.argv[1:]
    today = datetime.datetime(2026, 7, 31)
    if '--today' in args:
        i = args.index('--today'); today = datetime.datetime.strptime(args[i+1], '%Y-%m-%d'); del args[i:i+2]
    else:
        today = datetime.datetime.now()
    roster, stmts, exps, trips = load(today)
    if args and args[0] == '--snapshot':
        print(json.dumps(snapshot(roster, stmts, exps, trips), indent=1)); sys.exit()
    out = 'portal.html'
    if '--out' in args:
        i = args.index('--out'); out = args[i+1]; del args[i:i+2]
    who = ' '.join(args)
    open(out, 'w').write(render(who, roster, stmts, exps, trips, today))
    print('wrote', out)
